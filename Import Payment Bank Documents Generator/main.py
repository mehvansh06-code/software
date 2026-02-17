import customtkinter as ctk
import os
import sys
import logging
from datetime import datetime
from tkinter import messagebox
from PIL import Image
from typing import Dict, Any, Optional
from tkcalendar import DateEntry
import threading
import openpyxl  # <--- LIGHTWEIGHT REPLACEMENT FOR PANDAS

# =================================================
# 0. CONFIGURATION & ROBUST PATH FINDING
# =================================================

if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

LOG_DIR = os.path.join(APP_DIR, "logs")
OUTPUT_DIR = os.path.join(APP_DIR, "Generated_Docs")

if not os.path.exists(LOG_DIR): os.makedirs(LOG_DIR)
if not os.path.exists(OUTPUT_DIR): os.makedirs(OUTPUT_DIR)

logging.basicConfig(
    filename=os.path.join(LOG_DIR, f"app_log_{datetime.now().strftime('%Y_%m_%d')}.txt"),
    level=logging.INFO, 
    format='%(asctime)s - %(levelname)s - %(message)s'
)

ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

class Theme:
    BG = "#F9F9F9"
    SIDEBAR = "#FFFFFF"
    CARD = "#FFFFFF"
    BORDER = "#E0E0E0"
    TEXT_MAIN = "#333333"
    TEXT_SUB = "#666666"
    ACCENT = "#C5965E"
    BTN = "#004b8d"
    BTN_HOVER = "#003366"
    FONT = "Arial"

COMPANY_DB = {
    "Gujarat Flotex Pvt Ltd": {
        "filename": "ZHEJIANG FUSHENGDA.docx"
    },
    "GTEX Fabrics": {
        "filename": "Templategtex.docx"
    }
}

DOCUMENTS_LIST = [
    "REQUEST LETTER", "FEMA DECLARATION", 
    "OFAC DECLARATION", "FORM A1", 
    "INVOICE", "PACKING LIST", 
    "BILL OF LADING", "BANK ADVISE", 
    "COO", "BILL OF ENTRY", 
    "AIRWAY BILL", "INSURANCE COPY"
]

CURRENCY_NAMES = {
    "USD": "Dollars", "EUR": "Euro", "GBP": "Pounds", "JPY": "Japanese Yen", "CNY": "Chinese Yuan"
}

# =================================================
# 1. BUSINESS LOGIC LAYER
# =================================================
class DocumentService:
    
    @staticmethod
    def get_resource_path(filename):
        return os.path.join(APP_DIR, filename)

    @staticmethod
    def load_suppliers_background(callback):
        """Loads Excel using openpyxl (Fast & Light)"""
        def _task():
            try:
                file_path = DocumentService.get_resource_path("suppliers.xlsx")
                
                if not os.path.exists(file_path):
                    logging.error(f"Suppliers file missing at: {file_path}")
                    callback({}) 
                    return

                # LIGHTWEIGHT LOADING
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                
                # Get Headers
                headers = [cell.value for cell in ws[1]]
                if not headers or 'Supplier Name' not in headers:
                    logging.error("Excel missing 'Supplier Name' column")
                    callback({})
                    return

                # Map headers to indices
                header_map = {h.strip(): i for i, h in enumerate(headers) if h}
                sup_idx = header_map.get('Supplier Name')

                data = {}
                # Iterate rows (start from row 2)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[sup_idx]: continue # Skip empty names
                    
                    # Convert row tuple to dict based on headers
                    row_dict = {}
                    for col_name, idx in header_map.items():
                        val = row[idx] if idx < len(row) else ""
                        row_dict[col_name] = str(val) if val is not None else ""
                    
                    data[row_dict['Supplier Name'].strip()] = row_dict

                wb.close()
                callback(data)
                
            except Exception as e:
                logging.error(f"Excel Load Error: {e}")
                callback({})

        threading.Thread(target=_task, daemon=True).start()

    @staticmethod
    def validate_data(data: Dict[str, Any]) -> Optional[str]:
        fields = {
            'company_choice': "Importer (Company)",
            'invoice_no': "Invoice Number",
            'invoice_date': "Invoice Date",
            'shipment_date': "Shipment Date",
            'currency': "Currency",
            'raw_amount': "Remittance Amount",
            'beneficiary_name': "Supplier Name",
            'beneficiary_address': "Supplier Address",
            'beneficiary_country': "Supplier Country",
            'beneficiary_account': "Account Number",
            'bank_name': "Bank Name",
            'bank_swift': "SWIFT Code",
            'bank_address': "Bank Branch Address",
            'port_loading': "Port of Loading",
            'port_discharge': "Port of Discharge",
            'goods_desc': "Goods Description",
            'hsn_code': "HSN Code",
            'term': "IncoTerm",
            'mode_shipment': "Shipment Mode"
        }

        for key, label in fields.items():
            val = data.get(key, "")
            if not isinstance(val, str) or not val.strip():
                return f"MISSING: '{label}' is empty."
            if "Select" in val:
                return f"INVALID: Please select a valid '{label}'."

        qty = data.get('quantity', "").strip()
        parts = qty.split()
        if not parts or parts[0] == "":
            return "MISSING: 'Quantity' value is required."
        if "Select" in parts[1]:
            return "INVALID: Please select a valid Quantity Unit."

        try:
            if float(data['raw_amount']) <= 0:
                return "ERROR: Amount must be greater than 0."
        except ValueError:
            return "ERROR: Amount is not a valid number."

        return None

    @staticmethod
    def generate_document(template_name: str, context: Dict[str, Any]) -> str:
        from docxtpl import DocxTemplate
        
        template_path = DocumentService.get_resource_path(template_name)
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template not found at: {template_path}")

        now = datetime.now()
        save_dir = os.path.join(OUTPUT_DIR, str(now.year), now.strftime("%B"))
        if not os.path.exists(save_dir): os.makedirs(save_dir)

        base_name = f"{context['invoice_no']}_{context['currency']}"
        version = 1
        save_path = os.path.join(save_dir, f"{base_name}_v{version}.docx")
        while os.path.exists(save_path):
            version += 1
            save_path = os.path.join(save_dir, f"{base_name}_v{version}.docx")

        doc = DocxTemplate(template_path)
        doc.render(context)
        doc.save(save_path)
        return save_path

# =================================================
# 2. UI LAYER
# =================================================
class DocGeneratorApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Import Payment Doc Maker (Light)")
        self.geometry("1350x900")
        self.configure(fg_color=Theme.BG)
        
        self.supplier_map = {}
        self.supplier_list = ["Loading..."] 

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._setup_sidebar()
        self._setup_main_area()
        self._setup_form()
        
        self.after(100, self._start_background_load) 
        self.after(500, self._clear_dates)

    def _clear_dates(self):
        self.inv_date_picker.delete(0, "end")
        self.ship_date_picker.delete(0, "end")

    def _start_background_load(self):
        DocumentService.load_suppliers_background(self._on_suppliers_loaded)

    def _on_suppliers_loaded(self, data):
        self.supplier_map = data
        if not data:
            self.supplier_list = ["No Suppliers Found"]
            messagebox.showwarning("Data Missing", f"Could not load suppliers.xlsx.\nEnsure it is in:\n{APP_DIR}")
        else:
            self.supplier_list = list(self.supplier_map.keys())
        
        self.supplier_menu.configure(values=self.supplier_list)
        self.supplier_var.set("Select Supplier...")

    def _setup_sidebar(self):
        self.sidebar = ctk.CTkFrame(self, width=340, corner_radius=0, fg_color=Theme.SIDEBAR)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_propagate(False)

        ctk.CTkLabel(self.sidebar, text="CONTROLS", font=(Theme.FONT, 14, "bold"), text_color=Theme.TEXT_SUB).pack(pady=(30, 10), padx=25, anchor="w")

        self._create_label(self.sidebar, "Importer")
        self.company_var = ctk.StringVar(value="Select Company...")
        self.company_menu = ctk.CTkOptionMenu(self.sidebar, variable=self.company_var, values=list(COMPANY_DB.keys()), fg_color="white", button_color=Theme.ACCENT, text_color="black")
        self.company_menu.pack(padx=25, pady=(0, 20), fill="x")

        self._create_label(self.sidebar, "Select Supplier")
        self.supplier_var = ctk.StringVar(value="Loading...")
        self.supplier_menu = ctk.CTkOptionMenu(self.sidebar, variable=self.supplier_var, values=["Loading..."], command=self._auto_fill_supplier, fg_color="white", button_color="#555", text_color="black")
        self.supplier_menu.pack(padx=25, pady=(0, 20), fill="x")

        self._create_label(self.sidebar, "Payment Mode")
        self.pay_type_var = ctk.StringVar(value="Advance")
        ctk.CTkRadioButton(self.sidebar, text="Advance Payment", variable=self.pay_type_var, value="Advance", command=self._toggle_checklist, text_color="black", hover_color=Theme.ACCENT, fg_color=Theme.ACCENT).pack(pady=5, padx=25, anchor="w")
        ctk.CTkRadioButton(self.sidebar, text="Balance Payment", variable=self.pay_type_var, value="Balance", command=self._toggle_checklist, text_color="black", hover_color=Theme.ACCENT, fg_color=Theme.ACCENT).pack(pady=5, padx=25, anchor="w")

        ctk.CTkFrame(self.sidebar, height=1, fg_color="#ccc").pack(fill="x", pady=20, padx=25)

        self._create_label(self.sidebar, "Document Checklist")
        self.checklist_frame = ctk.CTkScrollableFrame(self.sidebar, height=350, fg_color="transparent")
        self.checklist_frame.pack(fill="x", padx=10, pady=(0, 20))
        
        self.check_vars = {}
        for doc in DOCUMENTS_LIST:
            is_checked = True if doc not in ["BANK ADVISE", "AIRWAY BILL", "INSURANCE COPY"] else False
            var = ctk.BooleanVar(value=is_checked)
            self.check_vars[doc] = var
            ctk.CTkCheckBox(self.checklist_frame, text=doc, variable=var, font=(Theme.FONT, 12), text_color="black", hover_color=Theme.ACCENT, fg_color=Theme.ACCENT, border_color="#888").pack(pady=4, anchor="w")

        self.part_adv_var = ctk.BooleanVar(value=False)
        self.part_adv_chk = ctk.CTkCheckBox(self.sidebar, text="Include BANK ADVISE", variable=self.part_adv_var, text_color="black", hover_color=Theme.ACCENT, fg_color=Theme.ACCENT)
        self._toggle_checklist()

    def _setup_main_area(self):
        self.main_area = ctk.CTkFrame(self, fg_color="transparent")
        self.main_area.grid(row=0, column=1, sticky="nsew", padx=30, pady=20)
        self.main_area.grid_rowconfigure(1, weight=1)
        self.main_area.grid_columnconfigure(0, weight=1)

        self.header = ctk.CTkFrame(self.main_area, height=80, fg_color="transparent")
        self.header.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        ctk.CTkLabel(self.header, text="Import Payment Doc Maker", font=(Theme.FONT, 26, "bold"), text_color="black").pack(side="left", anchor="center")
        
        try:
            logo_path = DocumentService.get_resource_path("GFPL LOGO R.png")
            if os.path.exists(logo_path):
                pil_image = Image.open(logo_path)
                ratio = 65 / pil_image.height
                new_width = int(pil_image.width * ratio)
                my_image = ctk.CTkImage(light_image=pil_image, dark_image=pil_image, size=(new_width, 65))
                ctk.CTkLabel(self.header, image=my_image, text="").pack(side="right", anchor="center")
            else:
                ctk.CTkLabel(self.header, text="GUJARAT FLOTEX", text_color=Theme.ACCENT, font=("Times New Roman", 20, "bold")).pack(side="right")
        except Exception as e:
            logging.error(f"Logo Load Error: {e}")

    def _setup_form(self):
        self.scroll_frame = ctk.CTkScrollableFrame(self.main_area, fg_color="transparent")
        self.scroll_frame.grid(row=1, column=0, sticky="nsew")
        self.scroll_frame.grid_columnconfigure(0, weight=1)

        # Card 1: Finance
        self.card_fin = self._create_card("💰 Remittance & Currency")
        self._create_label(self.card_fin, "Remittance Amount *", 1, 0)
        self._create_label(self.card_fin, "Currency *", 1, 1)
        self.amount_entry = ctk.CTkEntry(self.card_fin, height=35, border_color=Theme.BORDER, fg_color="white", text_color="black")
        self.amount_entry.grid(row=2, column=0, sticky="ew", padx=20, pady=(5, 15))
        
        self.currency_var = ctk.StringVar(value="Select...")
        self.currency_menu = ctk.CTkOptionMenu(self.card_fin, variable=self.currency_var, values=["USD", "EUR", "GBP", "JPY", "CNY"], fg_color="white", text_color="black", button_color="#ccc")
        self.currency_menu.grid(row=2, column=1, padx=20, pady=(5,15), sticky="ew")
        
        self._create_label(self.card_fin, "Total Invoice Value (If different)", 3, 0)
        self.inv_val_entry = ctk.CTkEntry(self.card_fin, height=35, border_color=Theme.BORDER, fg_color="white", text_color="black")
        self.inv_val_entry.grid(row=4, column=0, columnspan=2, padx=20, pady=(5,20), sticky="ew")

        # Card 2: Invoice
        self.card_inv = self._create_card("📄 Invoice Details")
        self.inv_no_entry = self._create_input(self.card_inv, "Invoice Number *", 1, 0)
        self._create_label(self.card_inv, "Invoice Date *", 1, 1)
        self.inv_date_picker = DateEntry(self.card_inv, width=12, background=Theme.BTN, foreground='white', borderwidth=2, date_pattern='dd-mm-yyyy')
        self.inv_date_picker.grid(row=2, column=1, padx=20, pady=15, sticky="ew")

        # Card 3: Supplier
        self.card_sup = self._create_card("🏭 Supplier Information")
        self.sup_name_entry = self._create_input_full(self.card_sup, "Supplier Name *", 1)
        self.sup_addr_entry = self._create_input_full(self.card_sup, "Address *", 3)
        self.sup_country_entry = self._create_input(self.card_sup, "Country *", 5, 0)
        self.sup_acc_entry = self._create_input(self.card_sup, "Account No *", 5, 1)
        self.sup_bank_entry = self._create_input(self.card_sup, "Bank Name *", 7, 0)
        self.sup_swift_entry = self._create_input(self.card_sup, "SWIFT Code *", 7, 1)
        self.sup_b_addr_entry = self._create_input_full(self.card_sup, "Bank Branch Address *", 9)

        # Card 4: Logistics
        self.card_log = self._create_card("🚢 Shipment & Goods")
        self._create_label(self.card_log, "IncoTerm *", 1, 0); self._create_label(self.card_log, "Mode *", 1, 1)
        
        self.term_var = ctk.StringVar(value="Select..."); ctk.CTkOptionMenu(self.card_log, variable=self.term_var, values=["CIF", "FOB", "EXW", "CFR"], fg_color="white", text_color="black", button_color="#ccc").grid(row=2, column=0, sticky="ew", padx=20, pady=(5,15))
        self.mode_var = ctk.StringVar(value="Select..."); ctk.CTkOptionMenu(self.card_log, variable=self.mode_var, values=["SEA", "AIR", "ROAD", "RAIL"], fg_color="white", text_color="black", button_color="#ccc").grid(row=2, column=1, sticky="ew", padx=20, pady=(5,15))
        
        self._create_label(self.card_log, "Shipment Date *", 3, 0); self._create_label(self.card_log, "Quantity & Unit *", 3, 1)
        self.ship_date_picker = DateEntry(self.card_log, width=12, background=Theme.BTN, foreground='white', borderwidth=2, date_pattern='dd-mm-yyyy')
        self.ship_date_picker.grid(row=4, column=0, padx=20, pady=15, sticky="ew")
        
        qty_frame = ctk.CTkFrame(self.card_log, fg_color="transparent")
        qty_frame.grid(row=4, column=1, sticky="ew", padx=20, pady=(5, 15))
        self.qty_val_entry = ctk.CTkEntry(qty_frame, height=35, width=120, border_color=Theme.BORDER, fg_color="white", text_color="black")
        self.qty_val_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
        
        self.qty_unit_var = ctk.StringVar(value="Select...")
        ctk.CTkOptionMenu(qty_frame, variable=self.qty_unit_var, width=90, values=["KGS", "MTR", "ROLLS", "PCS", "SETS"], fg_color="white", text_color="black", button_color="#ccc").pack(side="left")
        
        self.port_load_entry = self._create_input(self.card_log, "Port Loading *", 5, 0)
        self.port_discharge_entry = self._create_input(self.card_log, "Port Discharge *", 5, 1)
        self.goods_entry = self._create_input(self.card_log, "Goods Description *", 7, 0)
        self.hsn_entry = self._create_input(self.card_log, "HSN Code *", 7, 1)

        self.gen_btn = ctk.CTkButton(self.main_area, text="GENERATE DOCUMENTS", font=(Theme.FONT, 16, "bold"), height=50, fg_color=Theme.BTN, hover_color=Theme.BTN_HOVER, command=self._on_generate_click)
        self.gen_btn.grid(row=2, column=0, pady=20, sticky="ew")

    # --- UI HELPERS ---
    def _create_label(self, parent, text, r=None, c=None):
        label = ctk.CTkLabel(parent, text=text, font=(Theme.FONT, 12, "bold"), text_color=Theme.TEXT_SUB)
        if r is not None and c is not None: label.grid(row=r, column=c, sticky="w", padx=20 if parent != self.sidebar else 25)
        else: label.pack(padx=25, pady=(0,5), anchor="w")

    def _create_card(self, title):
        card = ctk.CTkFrame(self.scroll_frame, fg_color=Theme.CARD, corner_radius=10, border_width=1, border_color=Theme.BORDER)
        card.pack(fill="x", pady=10, padx=5)
        card.grid_columnconfigure((0,1), weight=1)
        ctk.CTkLabel(card, text=title, font=(Theme.FONT, 14, "bold"), text_color=Theme.BTN).grid(row=0, column=0, columnspan=2, sticky="w", padx=20, pady=(15, 10))
        return card

    def _create_input(self, parent, label, r, c):
        self._create_label(parent, label, r, c)
        entry = ctk.CTkEntry(parent, height=35, border_color=Theme.BORDER, fg_color="white", text_color="black")
        entry.grid(row=r+1, column=c, sticky="ew", padx=20, pady=(5, 15))
        return entry

    def _create_input_full(self, parent, label, r):
        self._create_label(parent, label, r, 0)
        entry = ctk.CTkEntry(parent, height=35, border_color=Theme.BORDER, fg_color="white", text_color="black")
        entry.grid(row=r+1, column=0, columnspan=2, sticky="ew", padx=20, pady=(5, 15))
        return entry

    def _toggle_checklist(self):
        if self.pay_type_var.get() == "Advance":
            self.checklist_frame.pack_forget()
            self.part_adv_chk.pack(pady=10, padx=25, anchor="w")
        else:
            self.part_adv_chk.pack_forget()
            self.checklist_frame.pack(fill="x", padx=10, pady=(0, 20))

    def _auto_fill_supplier(self, choice):
        if not self.supplier_map or choice not in self.supplier_map: return
        row = self.supplier_map.get(choice, {})
        fields = {
            self.sup_name_entry: 'Supplier Name', self.sup_addr_entry: 'Address',
            self.sup_country_entry: 'Country', self.sup_acc_entry: 'Account No',
            self.sup_bank_entry: 'Bank Name', self.sup_swift_entry: 'SWIFT',
            self.sup_b_addr_entry: 'Bank Address', self.goods_entry: 'Goods',
            self.hsn_entry: 'HSN'
        }
        for entry, col in fields.items():
            entry.delete(0, "end"); entry.insert(0, str(row.get(col, "")))

    def _collect_form_data(self) -> Dict[str, Any]:
        from num2words import num2words
        
        pay_type = self.pay_type_var.get()
        final_list = ["REQUEST LETTER", "FEMA DECLARATION", "OFAC DECLARATION", "ANNEXURE-B", "FORM A1", "INVOICE", "E-TRADE APPLICATION"] if pay_type == "Advance" else [doc for doc, var in self.check_vars.items() if var.get()]
        if pay_type == "Advance" and self.part_adv_var.get(): final_list.append("BANK ADVISE")
        
        raw_amt = self.amount_entry.get()
        try: amt_val = float(raw_amt)
        except: amt_val = 0.0
        
        curr_code = self.currency_var.get()
        curr_name = CURRENCY_NAMES.get(curr_code, curr_code)
        
        try: 
            if curr_code in ["JPY", "CNY"]:
                 amt_words = num2words(amt_val, lang='en').title() + " " + curr_name + " Only"
            else:
                 amt_words = num2words(amt_val, to='currency', currency=curr_code, lang='en').replace(f"{curr_code.lower()}, ", "").title()
        except: 
            amt_words = f"{amt_val} {curr_code}"

        inv_date_str = self.inv_date_picker.get()
        ship_date_str = self.ship_date_picker.get()

        return {
            'company_choice': self.company_var.get(),
            'date': datetime.now().strftime("%d-%m-%Y"),
            'invoice_no': self.inv_no_entry.get(),
            'invoice_date': inv_date_str,
            'shipment_date': ship_date_str,
            'currency': curr_code,
            'amount': f"{amt_val:,.2f}",
            'raw_amount': raw_amt,
            'amount_in_words': amt_words,
            'invoice_amount': self.inv_val_entry.get() or f"{amt_val:,.2f}",
            'quantity': f"{self.qty_val_entry.get()} {self.qty_unit_var.get()}",
            'beneficiary_name': self.sup_name_entry.get(),
            'beneficiary_address': self.sup_addr_entry.get(),
            'beneficiary_country': self.sup_country_entry.get(),
            'beneficiary_account': self.sup_acc_entry.get(),
            'bank_name': self.sup_bank_entry.get(),
            'bank_swift': self.sup_swift_entry.get(),
            'bank_address': self.sup_b_addr_entry.get(),
            'port_loading': self.port_load_entry.get(),
            'port_discharge': self.port_discharge_entry.get(),
            'purpose': f"PAYMENT FOR PURCHASE OF {self.goods_entry.get().upper()}",
            'goods_desc': self.goods_entry.get(),
            'hsn_code': self.hsn_entry.get(),
            'term': self.term_var.get(),
            'mode_shipment': self.mode_var.get(),
            'document_list': "".join([f"{i}.       {item}\n" for i, item in enumerate(final_list, 1)])
        }

    def _reset_form(self):
        self.company_var.set("Select Company...")
        self.supplier_var.set("Select Supplier...")
        self.currency_var.set("Select...")
        self.term_var.set("Select...")
        self.mode_var.set("Select...")
        self.qty_unit_var.set("Select...")
        self.pay_type_var.set("Advance")

        entries = [
            self.amount_entry, self.inv_val_entry, self.inv_no_entry,
            self.qty_val_entry, self.sup_name_entry, self.sup_addr_entry,
            self.sup_country_entry, self.sup_acc_entry, self.sup_bank_entry,
            self.sup_swift_entry, self.sup_b_addr_entry, self.port_load_entry,
            self.port_discharge_entry, self.goods_entry, self.hsn_entry
        ]
        for e in entries:
            e.delete(0, "end")
        
        self.part_adv_var.set(False)
        self._toggle_checklist()
        # Force clear with slight delay to override sticky widget behavior
        self.after(50, self._clear_dates)
        self.after(100, self._clear_dates) 

    def _on_generate_click(self):
        self.gen_btn.configure(state="disabled", text="PROCESSING...")
        self.update_idletasks()
        try:
            data = self._collect_form_data()
            error = DocumentService.validate_data(data)
            if error: 
                messagebox.showerror("Validation Error", error)
                return
            
            comp_data = COMPANY_DB[self.company_var.get()]
            save_path = DocumentService.generate_document(comp_data['filename'], data)
            
            messagebox.showinfo("Success", f"Document Created!\nSaved to:\n{save_path}")
            os.startfile(save_path)
            
            # Auto Reset after dialog closes
            self.after(500, self._reset_form)
            
        except Exception as e:
            logging.error(f"Generate Error: {e}")
            messagebox.showerror("System Error", f"Failed: {e}")
        finally:
            self.gen_btn.configure(state="normal", text="GENERATE DOCUMENTS")

if __name__ == "__main__":
    app = DocGeneratorApp()
    app.mainloop()